from openpyxl import load_workbook

def get_login_data(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append((row[0], row[1]))
    return data
